import tkinter as tk
from tkinter import filedialog, messagebox
import csv
import xml.etree.ElementTree as ET
import json
import os  # To handle file paths
import openpyxl  # For writing to Excel
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import Workbook

def get_columns_from_csv(file_path):
    with open(file_path, newline='') as csvfile:
        reader = csv.reader(csvfile)
        columns = next(reader)  # Read the header row to get the column names
        return columns

def get_data_from_xml(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    
    namespaces = {
        'saw': 'com.siebel.analytics.web/report/v1.1',
        'xsi': 'http://www.w3.org/2001/XMLSchema-instance',
        'sawx': 'com.siebel.analytics.web/expression/v1.1'
    }
    
    # Get the subjectArea from the saw:criteria
    dataSource= [root.find(".//saw:criteria", namespaces=namespaces).get('subjectArea').replace('"', '').lower()]
    
    # Get columns (expr elements for column)
    expr_elements_column = root.findall(".//saw:column//sawx:expr[@xsi:type='sawx:sqlExpression']", namespaces=namespaces)
    dataItem = [elem.text.replace('"', '').lower() for elem in expr_elements_column]

    # Get filters (expr elements for filter)
    expr_elements_filter = root.findall(".//saw:filter//sawx:expr[@xsi:type='sawx:sqlExpression']", namespaces=namespaces)
    dataFilter = [elem.text.replace('"', '').lower() for elem in expr_elements_filter]

    return dataItem, dataFilter, dataSource

def get_columns_from_json(file_path):
    with open(file_path) as jsonfile:
        data = json.load(jsonfile)
        columns = list(data[0].keys()) if data else []
        return columns

def get_data_from_file(file_path):
    if file_path.lower().endswith('.csv'):
        return get_columns_from_csv(file_path)
    elif file_path.lower().endswith('.xml'):
        return get_data_from_xml(file_path)
    elif file_path.lower().endswith('.json'):
        return get_columns_from_json(file_path)
    else:
        raise ValueError("Unsupported file format")

def calculate_matching_percentage(columns_file1, columns_file2):
    matching_columns = set(columns_file1).intersection(columns_file2)
    total_columns = len(set(columns_file1).union(columns_file2))
    
    if total_columns == 0:
        return 0
    
    matching_percentage = (len(matching_columns) / total_columns) * 100
    return matching_percentage

def append_to_excel(data_row, output_file, sheet_name):
    # Load the workbook or create a new one if it doesn't exist
    if os.path.exists(output_file):
        workbook = openpyxl.load_workbook(output_file)
    else:
        workbook = openpyxl.Workbook()
    
    # Remove the default sheet (if it exists) before proceeding
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # Check if the sheet already exists
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        # If the sheet doesn't exist, create it
        sheet = workbook.create_sheet(sheet_name)
        
        # Write main headers and subheaders for Sheet 1
        if sheet_name == "Sheet 1":
            main_headers = ["REPORT NAME", "REPORT PATH", "REPORT DataItem", "REPORT Datasource", "REPORT Data Filters"]
            subheaders = ["Name of the report", "Report Path", "Report DataItem", "Report DataSource", "Report Data Filters"]
            sheet.append(main_headers)  # Append main headers
            sheet.append(subheaders)    # Append subheaders
        
        # Write main headers and subheaders for Sheet 2
        elif sheet_name == "Sheet 2":
            main_headers = ["REPORT NAME", "DataItem Matched", "DataItem Match %", "Data Filters Matched", 
                            "Data Filters Match %", "Datasource Matched", "Datasource Match %", "Overall Match (%)"]
            subheaders = [
                "REPORTS COMBINED", 
                "DataItem Matched", 
                "Shows the percentage match between report data items", 
                "Data Filters Matched", 
                "Shows the percentage match between report data filters", 
                "Datasource Matched", 
                "Shows the percentage match between report Datasource", 
                "Shows the overall report matching percentage"
            ]
            sheet.append(main_headers)  # Append main headers
            sheet.append(subheaders)    # Append subheaders

        # Apply styles for main headers (bold, background color, and borders)
        main_header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
        main_header_font = Font(bold=True, color="000000")  # Bold and black text
        border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        
        # Apply to main headers (row 1)
        for cell in sheet[1]:  # The first row (main headers)
            cell.fill = main_header_fill
            cell.font = main_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align text
            cell.border = border  # Apply border to header cells

        # Apply styles for subheaders (italic, background color, and borders)
        subheader_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray background
        subheader_font = Font(italic=True, color="000000")  # Italic and black text
        for cell in sheet[2]:  # The second row (subheaders)
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align text
            cell.border = border  # Apply border to subheader cells

    # Write the data row to the sheet (append the data row)
    row_idx = sheet.max_row + 1  # Get the next empty row index
    for col_idx, value in enumerate(data_row, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=value)

        # Apply wrap text and center alignment for cells with line breaks
        if isinstance(value, str) and "\n" in value:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            # Apply center alignment for all cells (even without line breaks)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply border to all rows from the 3rd row onward
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(
                top=Side(border_style="thin", color="000000"),
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

    # Autofit column widths, but only consider the longest value among those with \n
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
        
        for cell in col:
            try:
                cell_value = str(cell.value)
                if '\n' in cell_value:
                    # Split by '\n' and get the maximum length of any individual line
                    max_length = max(max_length, max(len(line) for line in cell_value.split('\n')))
                else:
                    # Regular case for strings without '\n'
                    max_length = max(max_length, len(cell_value))
            except:
                pass
        
        # Adjust the column width by adding a little padding
        adjusted_width = max_length + 5  # Adding extra padding for header's appearance
        sheet.column_dimensions[column].width = adjusted_width

    # Increase row height for top row (main header) and subsequent rows
    sheet.row_dimensions[1].height = 40  # Increase height for header row (for more vertical space)
    sheet.row_dimensions[2].height = 35  # Increase height for subheader row

    workbook.save(output_file)


def compare_and_combine_columns(file_path1, file_path2):
    dataItem_file1, dataFilter_file1, dataSourcefile1 = get_data_from_file(file_path1)
    dataItem_file2, dataFilter_file2, dataSourcefile2 = get_data_from_file(file_path2)

    # Calculate matching percentages for dataItems (columns), dataFilters, and dataSources
    matching_percentage_dataItems = calculate_matching_percentage(dataItem_file1, dataItem_file2)
    matching_percentage_dataFilters = calculate_matching_percentage(dataFilter_file1, dataFilter_file2)
    matching_percentage_dataSources = calculate_matching_percentage(dataSourcefile1, dataSourcefile2)

    # Get the matched values for each category
    matched_dataItems = list(set(dataItem_file1).intersection(dataItem_file2))
    matched_dataFilters = list(set(dataFilter_file1).intersection(dataFilter_file2))
    matched_dataSources = list(set(dataSourcefile1).intersection(dataSourcefile2))

    # Calculate overall matching percentage as the average of the three
    overall_matching_percentage = (matching_percentage_dataItems + matching_percentage_dataFilters + matching_percentage_dataSources) / 3

    # If overall matching percentage is less than 75%, do not merge
    if overall_matching_percentage < 75:
        return None, overall_matching_percentage  # Return None if the match is below 75%

    # Prepare the report name and data row for Sheet 2
    file_name1 = os.path.splitext(os.path.basename(file_path1))[0]
    file_name2 = os.path.splitext(os.path.basename(file_path2))[0]
    report_name = f"{file_name1}_{file_name2}_COMBINED"
    
    # Format the matched values into strings with line breaks for DataItems, DataFilters, and DataSources
    matched_dataItems_str = ",\n".join(matched_dataItems) if matched_dataItems else "No Match"
    matched_dataFilters_str = ",\n".join(matched_dataFilters) if matched_dataFilters else "No Match"
    matched_dataSources_str = ",\n".join(matched_dataSources) if matched_dataSources else "No Match"

    # Prepare the data row for Sheet 1 for File 1
    report_path1 = file_path1  # Get the file path for the first report
    data_row_sheet1_file1 = [
        file_name1,
        report_path1,
        ",\n".join(dataItem_file1),  # Join DataItems with a line break
        ",\n".join(dataSourcefile1),  # Join Datasources with a line break
        ",\n".join(dataFilter_file1)  # Join DataFilters with a line break
    ]

    # Prepare the data row for Sheet 1 for File 2
    report_path2 = file_path2  # Get the file path for the second report
    data_row_sheet1_file2 = [
        file_name2,
        report_path2,
        ",\n".join(dataItem_file2),  # Join DataItems with a line break
        ",\n".join(dataSourcefile2),  # Join Datasources with a line break
        ",\n".join(dataFilter_file2)  # Join DataFilters with a line break
    ]

    # Prepare the data row for Sheet 2
    data_row_sheet2 = [
        report_name,
        matched_dataItems_str,
        matching_percentage_dataItems,
        matched_dataFilters_str,
        matching_percentage_dataFilters,
        matched_dataSources_str,
        matching_percentage_dataSources,
        overall_matching_percentage
    ]

    # Define the output file path based on both filenames
    output_file_path = f"{file_name1}_{file_name2}_merged.xlsx"

    # Append data to Sheet 1 (both File 1 and File 2) and Sheet 2 in the Excel file
    append_to_excel(data_row_sheet1_file1, output_file_path, sheet_name="Sheet 1")
    append_to_excel(data_row_sheet1_file2, output_file_path, sheet_name="Sheet 1")
    append_to_excel(data_row_sheet2, output_file_path, sheet_name="Sheet 2")

    return output_file_path, overall_matching_percentage


class ColumnComparerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Report Comparison Tool")
        self.root.config(bg="#f4f4f4")  # Soft gray background

        # Title Label
        self.title_label = tk.Label(root, text="Report Comparison Tool", font=("Helvetica", 18, "bold"), bg="#6E7F94", fg="white", padx=20, pady=20)
        self.title_label.grid(row=0, column=0, columnspan=3, pady=20, sticky="nsew")

        # Frames for File Selection
        self.frame = tk.Frame(root, bg="#f4f4f4")
        self.frame.grid(row=1, column=0, columnspan=3, padx=40, pady=15)

        # Select File 1
        self.label1 = tk.Label(self.frame, text="Select the first file:", font=("Arial", 12), bg="#f4f4f4")
        self.label1.grid(row=0, column=0, padx=10, pady=5, sticky="w")

        self.file1_button = tk.Button(self.frame, text="Browse", command=self.select_file1, font=("Arial", 10, "bold"), bg="#007BFF", fg="white", relief="flat")
        self.file1_button.grid(row=0, column=1, padx=10, pady=5)

        self.file1_label = tk.Label(self.frame, text="No file selected", font=("Arial", 10), bg="#f4f4f4", anchor="w", width=40)
        self.file1_label.grid(row=0, column=2, padx=10, pady=5, sticky="w")

        # Select File 2
        self.label2 = tk.Label(self.frame, text="Select the second file:", font=("Arial", 12), bg="#f4f4f4")
        self.label2.grid(row=1, column=0, padx=10, pady=5, sticky="w")

        self.file2_button = tk.Button(self.frame, text="Browse", command=self.select_file2, font=("Arial", 10, "bold"), bg="#007BFF", fg="white", relief="flat")
        self.file2_button.grid(row=1, column=1, padx=10, pady=5)

        self.file2_label = tk.Label(self.frame, text="No file selected", font=("Arial", 10), bg="#f4f4f4", anchor="w", width=40)
        self.file2_label.grid(row=1, column=2, padx=10, pady=5, sticky="w")

        # Compare Button
        self.compare_button = tk.Button(root, text="Compare", state=tk.DISABLED, command=self.compare_columns, font=("Arial", 12, "bold"), bg="#28a745", fg="white", relief="flat", padx=20, pady=10)
        self.compare_button.grid(row=2, column=0, columnspan=3, pady=20)

        # Result Label
        self.result_label = tk.Label(root, text="Results will be shown here", font=("Arial", 12), bg="#f4f4f4", anchor="w", height=4, relief="solid", padx=10, pady=10)
        self.result_label.grid(row=3, column=0, columnspan=3, padx=10, pady=10, sticky="ew")  # Stretch to fill the row

        # Initialize file paths
        self.file1 = None
        self.file2 = None

    def select_file1(self):
        self.file1 = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv"), ("XML Files", "*.xml"), ("JSON Files", "*.json")])
        if self.file1:
            self.file1_label.config(text=os.path.basename(self.file1))
        if self.file1 and self.file2:
            self.compare_button.config(state=tk.NORMAL)

    def select_file2(self):
        self.file2 = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv"), ("XML Files", "*.xml"), ("JSON Files", "*.json")])
        if self.file2:
            self.file2_label.config(text=os.path.basename(self.file2))
        if self.file1 and self.file2:
            self.compare_button.config(state=tk.NORMAL)

    def compare_columns(self):
        if not self.file1 or not self.file2:
            messagebox.showerror("Error", "Please select both files!")
            return
        
        try:
            output_file_path, matching_percentage = compare_and_combine_columns(self.file1, self.file2)
            result_message = f"Matching Percentage: {matching_percentage:.2f}%\n"
            result_message += f"Output File: {output_file_path}"
            self.result_label.config(text=result_message)
            messagebox.showinfo("Comparison Complete", f"Comparison complete. Results saved to {output_file_path}.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Running the Tkinter app
if __name__ == "__main__":
    root = tk.Tk()
    app = ColumnComparerApp(root)
    root.mainloop()